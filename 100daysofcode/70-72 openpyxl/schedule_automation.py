import datetime
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Alignment, Border, Side


# Playing with data in work schedule .xlsx

wb = load_workbook('shift_schedule1.xlsx')

ws1 = wb['Monday']
ws1['C2'] = datetime.date.today()
try:
    date_style1 = NamedStyle(name='datetime', number_format='DD.MM.YYYY')
    ws1['C2'].style = date_style1
except ValueError:
    pass
ws1['C2'].border = Border(bottom=Side(border_style="thin", color="000000"))

ws1['C2'].alignment = Alignment(horizontal='left')
ws1['C3'] = 'Sale'
ws1["M6"].value = '=COUNTIF(Monday[[#This Row],[7:00 AM]:[3:00 PM]],"*")'

ws1['B6'] = 'Michael'
ws1['B7'] = 'Bob'
ws1['B8'] = 'Miranda'
ws1['B9'] = 'Andrew'
ws1['B10'] = 'Jessica'

for row in range(2, ws1.max_row):
    ws1['C6'].value = 'manager'
    ws1['D6'].value = 'manager'
    ws1['E6'].value = 'manager'
    ws1['F6'].value = 'manager'
    ws1['G6'].value = 'manager'
    ws1['H6'].value = 'manager'
    ws1['I6'].value = 'dinner'
    ws1['I8'].value = 'dinner'
    ws1['I9'].value = 'dinner'

ws6 = wb['Saturday']

ws6['C2'] = datetime.date.today() + datetime.timedelta(days=5)
try:
    date_style6 = NamedStyle(name='datetime6', number_format='DD.MM.YYYY')
    ws6['C2'].style = date_style6
except ValueError:
    pass
ws6['C2'].border = Border(bottom=Side(border_style="thin", color="000000"))
ws6['C2'].alignment = Alignment(horizontal='left')

ws6['C3'] = 'Administration'
ws6['M6'].value = '=COUNTIF(Saturday[[#This Row],[7:00 AM]:[3:00 PM]],"*")'

if __name__ == '__main__':
    wb.save('shift_schedule1.xlsx')
